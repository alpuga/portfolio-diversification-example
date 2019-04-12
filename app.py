import openpyxl as xl
import math


def expected_returns(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(7, sheet.max_row + 1):
        expected_return_cell_one = sheet.cell(row, 3)
        expected_return_cell_two = sheet.cell(row, 5)
        expected_return_cell_three = sheet.cell(row, 7)

        expected_return = sheet.cell(row, 1).value * sheet['b2'].value + sheet.cell(row, 2).value * sheet['c2'].value

        expected_return_cell_one.value = expected_return
        expected_return_cell_two.value = expected_return
        expected_return_cell_three.value = expected_return

    wb.save(filename)


def standard_deviation_of_portfolios(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    for row in range(7, sheet.max_row + 1):
        sigma_cell_one = sheet.cell(row, 4)
        sigma_cell_two = sheet.cell(row, 6)
        sigma_cell_three = sheet.cell(row, 8)

        sigma_one = math.sqrt(
            (sheet.cell(row, 1).value ** 2) * (sheet['b3'].value ** 2) + (sheet.cell(row, 2).value ** 2) * (
                        sheet['c3'].value ** 2) + 2 * (sheet.cell(row, 1).value) * (sheet.cell(row, 2).value) * (
                sheet['d5'].value) * (sheet['b3'].value) * (sheet['c3'].value))

        sigma_two = math.sqrt(
            (sheet.cell(row, 1).value ** 2) * (sheet['b3'].value ** 2) + (sheet.cell(row, 2).value ** 2) * (
                        sheet['c3'].value ** 2) + 2 * (sheet.cell(row, 1).value) * (sheet.cell(row, 2).value) * (
                sheet['f5'].value) * (sheet['b3'].value) * (sheet['c3'].value))

        sigma_three = math.sqrt(
            (sheet.cell(row, 1).value ** 2) * (sheet['b3'].value ** 2) + (sheet.cell(row, 2).value ** 2) * (
                        sheet['c3'].value ** 2) + 2 * (sheet.cell(row, 1).value) * (sheet.cell(row, 2).value) * (
                sheet['h5'].value) * (sheet['b3'].value) * (sheet['c3'].value))

        sigma_cell_one.value = sigma_one
        sigma_cell_two.value = sigma_two
        sigma_cell_three.value = sigma_three

    wb.save(filename)
